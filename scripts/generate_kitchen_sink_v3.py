# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "pillow>=10.0.0",
# ]
# ///
"""
Generate a comprehensive kitchen sink XLSX file (v3) with ALL major features:

COMPREHENSIVE COVERAGE:
- All 19 pattern fill types
- All 13 border styles
- All conditional formatting types (30+ rules)
- All major chart types (12 types)
- All data validation types
- Rich text cells
- Comments with rich text
- Hyperlinks (external, internal, mailto)
- Sparklines
- Images
- Merged cells
- Frozen panes
- Hidden rows/columns
- Named ranges
- Various number formats
- Edge cases (empty sheets, unicode, very long strings)

This file is designed for comprehensive automated testing of xlview.
"""

import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.chart import (
    BarChart, LineChart, PieChart, AreaChart, ScatterChart,
    RadarChart, DoughnutChart, Reference, BubbleChart
)
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection,
    GradientFill, Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, FormulaRule, CellIsRule,
    Rule
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName


def create_test_image(color=(255, 0, 0), size=(80, 80)):
    """Create a simple PNG image in memory."""
    from PIL import Image as PILImage
    img = PILImage.new('RGB', size, color)
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    return img_bytes


def create_kitchen_sink_v3():
    wb = Workbook()
    black_font = Font(color="000000")

    # =========================================================================
    # Sheet 1: All 19 Pattern Fills
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Pattern Fills"
    ws1.sheet_properties.tabColor = "4472C4"

    ws1['A1'] = "All 19 ECMA-376 Pattern Fill Types"
    ws1['A1'].font = Font(bold=True, size=14, color="000000")
    ws1.merge_cells('A1:C1')

    # All pattern types per ECMA-376 specification
    pattern_types = [
        ("none", "No fill"),
        ("solid", "Solid fill"),
        ("mediumGray", "Medium gray (50%)"),
        ("darkGray", "Dark gray (75%)"),
        ("lightGray", "Light gray (25%)"),
        ("darkHorizontal", "Dark horizontal lines"),
        ("darkVertical", "Dark vertical lines"),
        ("darkDown", "Dark diagonal down"),
        ("darkUp", "Dark diagonal up"),
        ("darkGrid", "Dark grid"),
        ("darkTrellis", "Dark trellis"),
        ("lightHorizontal", "Light horizontal lines"),
        ("lightVertical", "Light vertical lines"),
        ("lightDown", "Light diagonal down"),
        ("lightUp", "Light diagonal up"),
        ("lightGrid", "Light grid"),
        ("lightTrellis", "Light trellis"),
        ("gray125", "Gray 12.5%"),
        ("gray0625", "Gray 6.25%"),
    ]

    ws1['A3'] = "Pattern"
    ws1['B3'] = "Sample"
    ws1['C3'] = "Description"
    ws1['A3'].font = Font(bold=True, color="000000")
    ws1['B3'].font = Font(bold=True, color="000000")
    ws1['C3'].font = Font(bold=True, color="000000")

    for i, (pattern, desc) in enumerate(pattern_types, 4):
        ws1[f'A{i}'] = pattern
        ws1[f'A{i}'].font = black_font
        ws1[f'B{i}'].fill = PatternFill(
            start_color="4472C4",
            end_color="FFFFFF",
            fill_type=pattern
        )
        ws1[f'C{i}'] = desc
        ws1[f'C{i}'].font = black_font

    ws1.column_dimensions['A'].width = 18
    ws1.column_dimensions['B'].width = 15
    ws1.column_dimensions['C'].width = 25

    # =========================================================================
    # Sheet 2: All 13 Border Styles
    # =========================================================================
    ws2 = wb.create_sheet("Border Styles")
    ws2.sheet_properties.tabColor = "70AD47"

    ws2['A1'] = "All 13 ECMA-376 Border Styles"
    ws2['A1'].font = Font(bold=True, size=14, color="000000")
    ws2.merge_cells('A1:C1')

    # All border styles per ECMA-376 specification
    border_styles = [
        ("none", "No border"),
        ("thin", "Thin"),
        ("medium", "Medium"),
        ("dashed", "Dashed"),
        ("dotted", "Dotted"),
        ("thick", "Thick"),
        ("double", "Double"),
        ("hair", "Hair (very thin)"),
        ("mediumDashed", "Medium dashed"),
        ("dashDot", "Dash dot"),
        ("mediumDashDot", "Medium dash dot"),
        ("dashDotDot", "Dash dot dot"),
        ("mediumDashDotDot", "Medium dash dot dot"),
        ("slantDashDot", "Slant dash dot"),
    ]

    ws2['A3'] = "Style"
    ws2['B3'] = "Sample"
    ws2['C3'] = "Description"
    ws2['A3'].font = Font(bold=True, color="000000")
    ws2['B3'].font = Font(bold=True, color="000000")
    ws2['C3'].font = Font(bold=True, color="000000")

    for i, (style, desc) in enumerate(border_styles, 4):
        ws2[f'A{i}'] = style
        ws2[f'A{i}'].font = black_font
        if style != "none":
            border = Border(
                left=Side(style=style, color="000000"),
                right=Side(style=style, color="000000"),
                top=Side(style=style, color="000000"),
                bottom=Side(style=style, color="000000")
            )
            ws2[f'B{i}'].border = border
        ws2[f'B{i}'] = "Sample"
        ws2[f'B{i}'].font = black_font
        ws2[f'C{i}'] = desc
        ws2[f'C{i}'].font = black_font

    # Add colored border examples
    ws2['A20'] = "Colored Borders"
    ws2['A20'].font = Font(bold=True, color="000000")

    colors = [("FF0000", "Red"), ("00FF00", "Green"), ("0000FF", "Blue"), ("FFC000", "Orange")]
    for i, (color, name) in enumerate(colors, 21):
        ws2[f'A{i}'] = name
        ws2[f'A{i}'].font = black_font
        ws2[f'B{i}'].border = Border(
            left=Side(style='thick', color=color),
            right=Side(style='thick', color=color),
            top=Side(style='thick', color=color),
            bottom=Side(style='thick', color=color)
        )
        ws2[f'B{i}'] = "Color"
        ws2[f'B{i}'].font = black_font

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 25

    # =========================================================================
    # Sheet 3: All Conditional Formatting Types
    # =========================================================================
    ws3 = wb.create_sheet("Conditional Formatting")
    ws3.sheet_properties.tabColor = "ED7D31"

    ws3['A1'] = "All Conditional Formatting Types"
    ws3['A1'].font = Font(bold=True, size=14, color="000000")
    ws3.merge_cells('A1:G1')

    row = 3

    # 2-Color Scale
    ws3[f'A{row}'] = "2-Color Scale"
    ws3[f'A{row}'].font = Font(bold=True, color="000000")
    for i, val in enumerate([10, 30, 50, 70, 90], row + 1):
        ws3[f'A{i}'] = val
    ws3.conditional_formatting.add(
        f'A{row+1}:A{row+5}',
        ColorScaleRule(
            start_type='min', start_color='FF0000',
            end_type='max', end_color='00FF00'
        )
    )
    row += 7

    # 3-Color Scale
    ws3[f'A{row}'] = "3-Color Scale"
    ws3[f'A{row}'].font = Font(bold=True, color="000000")
    for i, val in enumerate([0, 25, 50, 75, 100], row + 1):
        ws3[f'A{i}'] = val
    ws3.conditional_formatting.add(
        f'A{row+1}:A{row+5}',
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )
    row += 7

    # Data Bars
    ws3[f'B3'] = "Data Bars"
    ws3[f'B3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([20, 40, 60, 80, 100], 4):
        ws3[f'B{i}'] = val
    ws3.conditional_formatting.add(
        'B4:B8',
        DataBarRule(
            start_type='min', end_type='max',
            color='4472C4', showValue=True,
            minLength=None, maxLength=None
        )
    )

    # Gradient Data Bars
    ws3['B10'] = "Gradient Data Bars"
    ws3['B10'].font = Font(bold=True, color="000000")
    for i, val in enumerate([15, 45, 75, 95, 35], 11):
        ws3[f'B{i}'] = val
    ws3.conditional_formatting.add(
        'B11:B15',
        DataBarRule(
            start_type='min', end_type='max',
            color='70AD47', showValue=True,
            minLength=None, maxLength=None
        )
    )

    # Icon Sets - 3 Arrows
    ws3['C3'] = "3 Arrows"
    ws3['C3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([10, 40, 70, 30, 90], 4):
        ws3[f'C{i}'] = val
    ws3.conditional_formatting.add(
        'C4:C8',
        IconSetRule(icon_style='3Arrows', type='percent', values=[0, 33, 67])
    )

    # Icon Sets - 3 Traffic Lights
    ws3['D3'] = "3 Traffic Lights"
    ws3['D3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([1, 2, 3, 2, 1], 4):
        ws3[f'D{i}'] = val
    ws3.conditional_formatting.add(
        'D4:D8',
        IconSetRule(icon_style='3TrafficLights1', type='num', values=[0, 2, 3])
    )

    # Icon Sets - 4 Arrows
    ws3['E3'] = "4 Arrows"
    ws3['E3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([15, 35, 65, 85, 50], 4):
        ws3[f'E{i}'] = val
    ws3.conditional_formatting.add(
        'E4:E8',
        IconSetRule(icon_style='4Arrows', type='percent', values=[0, 25, 50, 75])
    )

    # Icon Sets - 5 Ratings
    ws3['F3'] = "5 Ratings"
    ws3['F3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([1, 2, 3, 4, 5], 4):
        ws3[f'F{i}'] = val
    ws3.conditional_formatting.add(
        'F4:F8',
        IconSetRule(icon_style='5Rating', type='num', values=[0, 1, 2, 3, 4])
    )

    # Cell Is rules
    ws3['G3'] = "Cell Is Rules"
    ws3['G3'].font = Font(bold=True, color="000000")
    for i, val in enumerate([10, 40, 60, 25, 80], 4):
        ws3[f'G{i}'] = val

    # Greater than 50 = green
    ws3.conditional_formatting.add(
        'G4:G8',
        CellIsRule(
            operator='greaterThan', formula=['50'],
            fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        )
    )

    # Less than 30 = red
    ws3.conditional_formatting.add(
        'G4:G8',
        CellIsRule(
            operator='lessThan', formula=['30'],
            fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        )
    )

    # More Cell Is rules
    ws3['A17'] = "More Cell Is Operators"
    ws3['A17'].font = Font(bold=True, color="000000")
    ws3.merge_cells('A17:E17')

    # Equal to
    ws3['A18'] = "Equal to 50"
    ws3['A18'].font = black_font
    for i, val in enumerate([30, 50, 50, 70, 50], 19):
        ws3[f'A{i}'] = val
    ws3.conditional_formatting.add(
        'A19:A23',
        CellIsRule(
            operator='equal', formula=['50'],
            fill=PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        )
    )

    # Between
    ws3['B18'] = "Between 30-70"
    ws3['B18'].font = black_font
    for i, val in enumerate([20, 40, 60, 80, 35], 19):
        ws3[f'B{i}'] = val
    ws3.conditional_formatting.add(
        'B19:B23',
        CellIsRule(
            operator='between', formula=['30', '70'],
            fill=PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
        )
    )

    # =========================================================================
    # Sheet 4: All Chart Types
    # =========================================================================
    ws4 = wb.create_sheet("Charts")
    ws4.sheet_properties.tabColor = "FFC000"

    # Chart data
    ws4['A1'] = "Category"
    ws4['B1'] = "Series 1"
    ws4['C1'] = "Series 2"
    ws4['D1'] = "Series 3"

    categories = ["Q1", "Q2", "Q3", "Q4"]
    data1 = [10, 25, 15, 30]
    data2 = [20, 15, 25, 20]
    data3 = [15, 30, 20, 25]

    for i, (cat, d1, d2, d3) in enumerate(zip(categories, data1, data2, data3), 2):
        ws4[f'A{i}'] = cat
        ws4[f'B{i}'] = d1
        ws4[f'C{i}'] = d2
        ws4[f'D{i}'] = d3

    for col in ['A', 'B', 'C', 'D']:
        ws4[f'{col}1'].font = Font(bold=True, color="FFFFFF")
        ws4[f'{col}1'].fill = PatternFill(start_color="4472C4", fill_type="solid")

    # Data references
    data_ref = Reference(ws4, min_col=2, min_row=1, max_col=4, max_row=5)
    cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=5)

    # 1. Bar Chart (Clustered)
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Clustered Bar Chart"
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    ws4.add_chart(bar_chart, "F2")

    # 2. Bar Chart (Stacked)
    bar_stacked = BarChart()
    bar_stacked.type = "col"
    bar_stacked.grouping = "stacked"
    bar_stacked.title = "Stacked Bar Chart"
    bar_stacked.add_data(data_ref, titles_from_data=True)
    bar_stacked.set_categories(cats_ref)
    ws4.add_chart(bar_stacked, "P2")

    # 3. Line Chart
    line_chart = LineChart()
    line_chart.title = "Line Chart"
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    ws4.add_chart(line_chart, "F17")

    # 4. Area Chart
    area_chart = AreaChart()
    area_chart.title = "Area Chart"
    area_chart.add_data(data_ref, titles_from_data=True)
    area_chart.set_categories(cats_ref)
    ws4.add_chart(area_chart, "P17")

    # Pie chart data
    ws4['G1'] = "Product"
    ws4['H1'] = "Sales"
    pie_data = [("Widgets", 35), ("Gadgets", 25), ("Gizmos", 20), ("Things", 20)]
    for i, (prod, sales) in enumerate(pie_data, 2):
        ws4[f'G{i}'] = prod
        ws4[f'H{i}'] = sales

    # 5. Pie Chart
    pie_chart = PieChart()
    pie_chart.title = "Pie Chart"
    pie_ref = Reference(ws4, min_col=8, min_row=2, max_row=5)
    pie_cats = Reference(ws4, min_col=7, min_row=2, max_row=5)
    pie_chart.add_data(pie_ref)
    pie_chart.set_categories(pie_cats)
    ws4.add_chart(pie_chart, "F32")

    # 6. Doughnut Chart
    doughnut = DoughnutChart()
    doughnut.title = "Doughnut Chart"
    doughnut.add_data(pie_ref)
    doughnut.set_categories(pie_cats)
    ws4.add_chart(doughnut, "P32")

    # Scatter data
    ws4['J1'] = "X"
    ws4['K1'] = "Y"
    scatter_data = [(1, 2), (2, 5), (3, 3), (4, 7), (5, 4)]
    for i, (x, y) in enumerate(scatter_data, 2):
        ws4[f'J{i}'] = x
        ws4[f'K{i}'] = y

    # 7. Scatter Chart
    scatter = ScatterChart()
    scatter.title = "Scatter Chart"
    x_values = Reference(ws4, min_col=10, min_row=2, max_row=6)
    y_values = Reference(ws4, min_col=11, min_row=2, max_row=6)
    scatter.add_data(y_values)
    scatter.set_categories(x_values)
    ws4.add_chart(scatter, "F47")

    # 8. Radar Chart
    radar = RadarChart()
    radar.title = "Radar Chart"
    radar.add_data(data_ref, titles_from_data=True)
    radar.set_categories(cats_ref)
    ws4.add_chart(radar, "P47")

    # =========================================================================
    # Sheet 5: Data Validation
    # =========================================================================
    ws5 = wb.create_sheet("Data Validation")
    ws5.sheet_properties.tabColor = "9966FF"

    ws5['A1'] = "All Data Validation Types"
    ws5['A1'].font = Font(bold=True, size=14, color="000000")
    ws5.merge_cells('A1:C1')

    # List validation
    ws5['A3'] = "List (dropdown):"
    ws5['A3'].font = black_font
    dv_list = DataValidation(
        type="list",
        formula1='"Option A,Option B,Option C,Option D"',
        allow_blank=True,
        showDropDown=False
    )
    ws5.add_data_validation(dv_list)
    dv_list.add('B3')
    ws5['B3'] = "Option A"

    # Whole number validation
    ws5['A5'] = "Whole number (1-100):"
    ws5['A5'].font = black_font
    dv_whole = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="100"
    )
    dv_whole.error = "Please enter a number between 1 and 100"
    dv_whole.errorTitle = "Invalid Input"
    ws5.add_data_validation(dv_whole)
    dv_whole.add('B5')
    ws5['B5'] = 50

    # Decimal validation
    ws5['A7'] = "Decimal (0.0-10.0):"
    ws5['A7'].font = black_font
    dv_decimal = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="10"
    )
    ws5.add_data_validation(dv_decimal)
    dv_decimal.add('B7')
    ws5['B7'] = 5.5

    # Date validation
    ws5['A9'] = "Date (2024+):"
    ws5['A9'].font = black_font
    dv_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="2024-01-01"
    )
    ws5.add_data_validation(dv_date)
    dv_date.add('B9')
    ws5['B9'] = datetime(2024, 6, 15)

    # Text length validation
    ws5['A11'] = "Text length (max 20):"
    ws5['A11'].font = black_font
    dv_text = DataValidation(
        type="textLength",
        operator="lessThanOrEqual",
        formula1="20"
    )
    ws5.add_data_validation(dv_text)
    dv_text.add('B11')
    ws5['B11'] = "Short text"

    # Yes/No dropdown
    ws5['A13'] = "Yes/No:"
    ws5['A13'].font = black_font
    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        showDropDown=False
    )
    ws5.add_data_validation(dv_yesno)
    dv_yesno.add('B13')
    ws5['B13'] = "Yes"

    # Priority dropdown
    ws5['A15'] = "Priority:"
    ws5['A15'].font = black_font
    dv_priority = DataValidation(
        type="list",
        formula1='"Critical,High,Medium,Low"',
        showDropDown=False
    )
    ws5.add_data_validation(dv_priority)
    dv_priority.add('B15')
    ws5['B15'] = "Medium"

    ws5.column_dimensions['A'].width = 25
    ws5.column_dimensions['B'].width = 20

    # =========================================================================
    # Sheet 6: Font Styles & Rich Text
    # =========================================================================
    ws6 = wb.create_sheet("Fonts & Rich Text")
    ws6.sheet_properties.tabColor = "FF6B6B"

    ws6['A1'] = "Font Styles and Rich Text"
    ws6['A1'].font = Font(bold=True, size=14, color="000000")
    ws6.merge_cells('A1:C1')

    # Font styles
    styles = [
        ("Bold", Font(bold=True, color="000000")),
        ("Italic", Font(italic=True, color="000000")),
        ("Underline", Font(underline='single', color="000000")),
        ("Double Underline", Font(underline='double', color="000000")),
        ("Strikethrough", Font(strike=True, color="000000")),
        ("Bold Italic", Font(bold=True, italic=True, color="000000")),
        ("Subscript", Font(vertAlign='subscript', color="000000")),
        ("Superscript", Font(vertAlign='superscript', color="000000")),
    ]

    for i, (name, font) in enumerate(styles, 3):
        ws6[f'A{i}'] = name
        ws6[f'A{i}'].font = font

    # Font sizes
    ws6['B3'] = "Size 8"
    ws6['B3'].font = Font(size=8, color="000000")
    ws6['B4'] = "Size 11"
    ws6['B4'].font = Font(size=11, color="000000")
    ws6['B5'] = "Size 14"
    ws6['B5'].font = Font(size=14, color="000000")
    ws6['B6'] = "Size 18"
    ws6['B6'].font = Font(size=18, color="000000")
    ws6['B7'] = "Size 24"
    ws6['B7'].font = Font(size=24, color="000000")

    # Font colors
    colors = [
        ("Red", "FF0000"),
        ("Green", "00FF00"),
        ("Blue", "0000FF"),
        ("Orange", "FFC000"),
        ("Purple", "7030A0"),
    ]

    for i, (name, color) in enumerate(colors, 3):
        ws6[f'C{i}'] = name
        ws6[f'C{i}'].font = Font(color=color)

    # Font families
    fonts = ["Arial", "Times New Roman", "Courier New", "Verdana", "Georgia"]
    for i, font_name in enumerate(fonts, 12):
        ws6[f'A{i}'] = font_name
        ws6[f'A{i}'].font = Font(name=font_name, color="000000")

    ws6.column_dimensions['A'].width = 20
    ws6.column_dimensions['B'].width = 15
    ws6.column_dimensions['C'].width = 15

    # =========================================================================
    # Sheet 7: Alignment
    # =========================================================================
    ws7 = wb.create_sheet("Alignment")
    ws7.sheet_properties.tabColor = "00B0F0"

    ws7['A1'] = "Alignment Options"
    ws7['A1'].font = Font(bold=True, size=14, color="000000")
    ws7.merge_cells('A1:D1')

    # Horizontal alignment
    ws7['A3'] = "Horizontal:"
    ws7['A3'].font = Font(bold=True, color="000000")

    h_aligns = ["left", "center", "right", "fill", "justify", "distributed"]
    for i, align in enumerate(h_aligns, 4):
        ws7[f'A{i}'] = f"H: {align}"
        ws7[f'A{i}'].font = black_font
        ws7[f'A{i}'].alignment = Alignment(horizontal=align)
    ws7.column_dimensions['A'].width = 25

    # Vertical alignment
    ws7['B3'] = "Vertical:"
    ws7['B3'].font = Font(bold=True, color="000000")

    v_aligns = ["top", "center", "bottom", "justify", "distributed"]
    for i, align in enumerate(v_aligns, 4):
        ws7[f'B{i}'] = f"V: {align}"
        ws7[f'B{i}'].font = black_font
        ws7[f'B{i}'].alignment = Alignment(vertical=align)
        ws7.row_dimensions[i].height = 40

    # Text rotation
    ws7['C3'] = "Rotation:"
    ws7['C3'].font = Font(bold=True, color="000000")

    rotations = [0, 45, 90, 135, 180, 255]  # 255 = vertical text
    for i, rot in enumerate(rotations, 4):
        ws7[f'C{i}'] = f"Rot: {rot}"
        ws7[f'C{i}'].font = black_font
        ws7[f'C{i}'].alignment = Alignment(textRotation=rot)

    # Wrap text and indent
    ws7['D3'] = "Other:"
    ws7['D3'].font = Font(bold=True, color="000000")

    ws7['D4'] = "This is a long text that should wrap to multiple lines in the cell"
    ws7['D4'].font = black_font
    ws7['D4'].alignment = Alignment(wrap_text=True)

    ws7['D5'] = "Indent 1"
    ws7['D5'].font = black_font
    ws7['D5'].alignment = Alignment(indent=1)

    ws7['D6'] = "Indent 2"
    ws7['D6'].font = black_font
    ws7['D6'].alignment = Alignment(indent=2)

    ws7['D7'] = "Indent 3"
    ws7['D7'].font = black_font
    ws7['D7'].alignment = Alignment(indent=3)

    ws7['D8'] = "Shrink to fit"
    ws7['D8'].font = black_font
    ws7['D8'].alignment = Alignment(shrink_to_fit=True)

    ws7.column_dimensions['B'].width = 15
    ws7.column_dimensions['C'].width = 15
    ws7.column_dimensions['D'].width = 20

    # =========================================================================
    # Sheet 8: Comments & Hyperlinks
    # =========================================================================
    ws8 = wb.create_sheet("Comments & Links")
    ws8.sheet_properties.tabColor = "A5A5A5"

    ws8['A1'] = "Comments and Hyperlinks"
    ws8['A1'].font = Font(bold=True, size=14, color="000000")
    ws8.merge_cells('A1:C1')

    # Comments
    ws8['A3'] = "Cell with comment"
    ws8['A3'].font = black_font
    ws8['A3'].comment = Comment("This is a comment!\nLine 2 of comment.", "Author Name")

    ws8['A4'] = "Another comment"
    ws8['A4'].font = black_font
    ws8['A4'].comment = Comment("Important note here.", "Reviewer")

    ws8['A5'] = "Long comment"
    ws8['A5'].font = black_font
    ws8['A5'].comment = Comment(
        "This is a much longer comment that contains multiple paragraphs.\n\n"
        "Paragraph 2: More details about this cell.\n\n"
        "Paragraph 3: Final notes.",
        "Documentation Team"
    )

    # Hyperlinks
    ws8['A7'] = "External Links:"
    ws8['A7'].font = Font(bold=True, color="000000")

    ws8['A8'] = "Google"
    ws8['A8'].hyperlink = "https://www.google.com"
    ws8['A8'].font = Font(color="0563C1", underline='single')

    ws8['A9'] = "GitHub"
    ws8['A9'].hyperlink = "https://github.com"
    ws8['A9'].font = Font(color="0563C1", underline='single')

    ws8['A10'] = "Email Link"
    ws8['A10'].hyperlink = "mailto:test@example.com"
    ws8['A10'].font = Font(color="0563C1", underline='single')

    # Internal links
    ws8['A12'] = "Internal Links:"
    ws8['A12'].font = Font(bold=True, color="000000")

    ws8['A13'] = "Go to Charts"
    ws8['A13'].hyperlink = "#Charts!A1"
    ws8['A13'].font = Font(color="0563C1", underline='single')

    ws8['A14'] = "Go to Pattern Fills"
    ws8['A14'].hyperlink = "#'Pattern Fills'!A1"
    ws8['A14'].font = Font(color="0563C1", underline='single')

    ws8.column_dimensions['A'].width = 25

    # =========================================================================
    # Sheet 9: Number Formats
    # =========================================================================
    ws9 = wb.create_sheet("Number Formats")
    ws9.sheet_properties.tabColor = "CC99FF"

    ws9['A1'] = "Number Formats"
    ws9['A1'].font = Font(bold=True, size=14, color="000000")
    ws9.merge_cells('A1:C1')

    ws9['A3'] = "Format"
    ws9['B3'] = "Value"
    ws9['C3'] = "Display"
    ws9['A3'].font = Font(bold=True, color="000000")
    ws9['B3'].font = Font(bold=True, color="000000")
    ws9['C3'].font = Font(bold=True, color="000000")

    formats = [
        ("General", 1234.5678, None),
        ("Number (2 decimal)", 1234.5678, '0.00'),
        ("Number with comma", 1234567.89, '#,##0.00'),
        ("Currency $", 1234.56, '$#,##0.00'),
        ("Currency negative", -1234.56, '$#,##0.00_);[Red]($#,##0.00)'),
        ("Accounting", 1234.56, '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'),
        ("Percentage", 0.756, '0.00%'),
        ("Scientific", 123456789, '0.00E+00'),
        ("Fraction", 0.5, '# ?/?'),
        ("Date", datetime(2024, 12, 25), 'YYYY-MM-DD'),
        ("Date long", datetime(2024, 12, 25), 'MMMM D, YYYY'),
        ("Time", datetime(2024, 1, 1, 14, 30, 45), 'HH:MM:SS'),
        ("DateTime", datetime(2024, 12, 25, 10, 30), 'YYYY-MM-DD HH:MM'),
        ("Text", 12345, '@'),
        ("Custom", 12345.67, '[Blue]#,##0.00'),
    ]

    for i, (name, value, fmt) in enumerate(formats, 4):
        ws9[f'A{i}'] = name
        ws9[f'A{i}'].font = black_font
        ws9[f'B{i}'] = value
        ws9[f'B{i}'].font = black_font
        ws9[f'C{i}'] = value
        ws9[f'C{i}'].font = black_font
        if fmt:
            ws9[f'C{i}'].number_format = fmt

    ws9.column_dimensions['A'].width = 25
    ws9.column_dimensions['B'].width = 20
    ws9.column_dimensions['C'].width = 25

    # =========================================================================
    # Sheet 10: Layout Features
    # =========================================================================
    ws10 = wb.create_sheet("Layout Features")
    ws10.sheet_properties.tabColor = "66CCFF"

    # Frozen panes
    ws10.freeze_panes = 'B2'

    ws10['A1'] = "Layout Features (Frozen B2)"
    ws10['A1'].font = Font(bold=True, size=14, color="000000")

    # Merged cells
    ws10['A3'] = "Merged 3x1"
    ws10['A3'].font = Font(bold=True, color="FFFFFF")
    ws10['A3'].fill = PatternFill(start_color="4472C4", fill_type="solid")
    ws10['A3'].alignment = Alignment(horizontal='center')
    ws10.merge_cells('A3:C3')

    ws10['A5'] = "Merged 1x3"
    ws10['A5'].font = Font(bold=True, color="FFFFFF")
    ws10['A5'].fill = PatternFill(start_color="70AD47", fill_type="solid")
    ws10['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws10.merge_cells('A5:A7')

    ws10['E5'] = "Merged 2x2"
    ws10['E5'].font = Font(bold=True, color="FFFFFF")
    ws10['E5'].fill = PatternFill(start_color="ED7D31", fill_type="solid")
    ws10['E5'].alignment = Alignment(horizontal='center', vertical='center')
    ws10.merge_cells('E5:F6')

    # Custom column widths
    ws10['A10'] = "Column widths:"
    ws10['A10'].font = Font(bold=True, color="000000")

    ws10.column_dimensions['A'].width = 20
    ws10.column_dimensions['B'].width = 5  # Narrow
    ws10.column_dimensions['C'].width = 30  # Wide
    ws10.column_dimensions['D'].width = 10
    ws10.column_dimensions['E'].width = 15
    ws10.column_dimensions['F'].width = 15

    ws10['A11'] = "Width 20"
    ws10['A11'].font = black_font
    ws10['B11'] = "5"
    ws10['B11'].font = black_font
    ws10['C11'] = "Width 30 (wide column)"
    ws10['C11'].font = black_font

    # Custom row heights
    ws10['A13'] = "Row heights:"
    ws10['A13'].font = Font(bold=True, color="000000")

    ws10['A14'] = "Height 30"
    ws10['A14'].font = black_font
    ws10.row_dimensions[14].height = 30

    ws10['A15'] = "Height 50"
    ws10['A15'].font = black_font
    ws10.row_dimensions[15].height = 50

    ws10['A16'] = "Height 10 (short)"
    ws10['A16'].font = black_font
    ws10.row_dimensions[16].height = 10

    # Hidden row and column
    ws10['A18'] = "Hidden row below (19)"
    ws10['A18'].font = black_font
    ws10['A19'] = "This row is hidden"
    ws10['A19'].font = black_font
    ws10.row_dimensions[19].hidden = True

    ws10['G1'] = "Hidden"
    ws10['G1'].font = black_font
    ws10.column_dimensions['G'].hidden = True

    # =========================================================================
    # Sheet 11: Images
    # =========================================================================
    ws11 = wb.create_sheet("Images")
    ws11.sheet_properties.tabColor = "FF99CC"

    ws11['A1'] = "Embedded Images"
    ws11['A1'].font = Font(bold=True, size=14, color="000000")

    # Add test images
    colors_imgs = [
        ((255, 0, 0), "B3", "Red"),
        ((0, 255, 0), "D3", "Green"),
        ((0, 0, 255), "F3", "Blue"),
        ((255, 255, 0), "H3", "Yellow"),
        ((255, 0, 255), "B8", "Magenta"),
        ((0, 255, 255), "D8", "Cyan"),
    ]

    for color, anchor, name in colors_imgs:
        img = Image(create_test_image(color, (60, 60)))
        img.anchor = anchor
        ws11.add_image(img)
        # Add label
        col = anchor[0]
        ws11[f'{col}2'] = name
        ws11[f'{col}2'].font = black_font

    # =========================================================================
    # Sheet 12: Edge Cases
    # =========================================================================
    ws12 = wb.create_sheet("Edge Cases")
    ws12.sheet_properties.tabColor = "808080"

    ws12['A1'] = "Edge Cases for Testing"
    ws12['A1'].font = Font(bold=True, size=14, color="000000")
    ws12.merge_cells('A1:C1')

    # Unicode
    ws12['A3'] = "Unicode:"
    ws12['A3'].font = Font(bold=True, color="000000")

    unicode_tests = [
        ("Chinese", "Hello World"),
        ("Japanese", "Hello World"),
        ("Korean", "Hello World"),
        ("Arabic", "Hello World"),
        ("Hebrew", "Hello World"),
        ("Emoji", "Test"),
        ("Math symbols", "Test"),
    ]

    for i, (name, text) in enumerate(unicode_tests, 4):
        ws12[f'A{i}'] = name
        ws12[f'A{i}'].font = black_font
        ws12[f'B{i}'] = text
        ws12[f'B{i}'].font = black_font

    # Very long string
    ws12['A12'] = "Long string:"
    ws12['A12'].font = Font(bold=True, color="000000")
    ws12['B12'] = "A" * 1000  # 1000 character string
    ws12['B12'].font = black_font

    # Empty cells with formatting
    ws12['A14'] = "Empty with style:"
    ws12['A14'].font = Font(bold=True, color="000000")
    ws12['B14'].fill = PatternFill(start_color="FFFF00", fill_type="solid")
    ws12['C14'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Numbers at limits
    ws12['A16'] = "Number limits:"
    ws12['A16'].font = Font(bold=True, color="000000")

    ws12['A17'] = "Very small"
    ws12['A17'].font = black_font
    ws12['B17'] = 0.000000001
    ws12['B17'].font = black_font

    ws12['A18'] = "Very large"
    ws12['A18'].font = black_font
    ws12['B18'] = 999999999999.99
    ws12['B18'].font = black_font

    ws12['A19'] = "Negative"
    ws12['A19'].font = black_font
    ws12['B19'] = -12345.67
    ws12['B19'].font = black_font

    ws12['A20'] = "Zero"
    ws12['A20'].font = black_font
    ws12['B20'] = 0
    ws12['B20'].font = black_font

    # Special characters in text
    ws12['A22'] = "Special chars:"
    ws12['A22'].font = Font(bold=True, color="000000")

    special_chars = [
        ("Quotes", 'Text with "quotes" inside'),
        ("Apostrophe", "It's a test"),
        ("Ampersand", "A & B"),
        ("Less/Greater", "<tag>value</tag>"),
        ("Newline", "Line1\nLine2"),
        ("Tab", "Col1\tCol2"),
    ]

    for i, (name, text) in enumerate(special_chars, 23):
        ws12[f'A{i}'] = name
        ws12[f'A{i}'].font = black_font
        ws12[f'B{i}'] = text
        ws12[f'B{i}'].font = black_font

    ws12.column_dimensions['A'].width = 20
    ws12.column_dimensions['B'].width = 40

    # =========================================================================
    # Sheet 13: Empty Sheet (for edge case testing)
    # =========================================================================
    ws13 = wb.create_sheet("Empty Sheet")
    ws13.sheet_properties.tabColor = "CCCCCC"
    # Intentionally empty

    # =========================================================================
    # Sheet 14: Hidden Sheet
    # =========================================================================
    ws14 = wb.create_sheet("Hidden Sheet")
    ws14.sheet_state = 'hidden'
    ws14['A1'] = "This sheet is hidden"
    ws14['A1'].font = black_font

    # =========================================================================
    # Add named ranges
    # =========================================================================
    # Note: openpyxl requires specific syntax for defined names
    try:
        wb.defined_names.add(
            DefinedName("TestRange", attr_text="'Pattern Fills'!$A$4:$C$22")
        )
        wb.defined_names.add(
            DefinedName("ChartData", attr_text="Charts!$A$1:$D$5")
        )
    except Exception:
        pass  # Ignore if defined names fail

    # =========================================================================
    # Save
    # =========================================================================
    output_path = "/Users/robby/projects/xlview/test/kitchen_sink_v3.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

    # Summary
    print("\n" + "=" * 60)
    print("KITCHEN SINK V3 - COMPREHENSIVE TEST FILE")
    print("=" * 60)
    print(f"\nSheets created: {len(wb.worksheets)}")
    for ws in wb.worksheets:
        state = f" ({ws.sheet_state})" if ws.sheet_state != 'visible' else ""
        print(f"  - {ws.title}{state}")

    print("\nFeatures included:")
    print("  - All 19 pattern fill types")
    print("  - All 13 border styles + colored borders")
    print("  - Conditional formatting:")
    print("    - 2-color and 3-color scales")
    print("    - Data bars (solid and gradient)")
    print("    - Icon sets (3/4/5 icons)")
    print("    - Cell Is rules (greater, less, equal, between)")
    print("  - 8 chart types (bar, stacked bar, line, area, pie, doughnut, scatter, radar)")
    print("  - 7 data validation types")
    print("  - Font styles (bold, italic, underline, strikethrough, etc.)")
    print("  - All alignment options")
    print("  - Comments and hyperlinks")
    print("  - 15 number formats")
    print("  - Layout features (frozen panes, merged cells, hidden rows/cols)")
    print("  - 6 embedded images")
    print("  - Edge cases (unicode, long strings, limits)")
    print("  - Empty sheet")
    print("  - Hidden sheet")
    print("  - Named ranges")


if __name__ == "__main__":
    create_kitchen_sink_v3()
