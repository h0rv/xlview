#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
# ]
# ///
"""
Generate test Excel files for xlview testing.

Usage:
    uv run scripts/generate_test_xlsx.py                    # Generate all test files
    uv run scripts/generate_test_xlsx.py --only kitchen     # Just kitchen sink
    uv run scripts/generate_test_xlsx.py --only large       # Just large file
    uv run scripts/generate_test_xlsx.py --rows 1000        # Custom large file size
"""

import argparse
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, GradientFill, Border, Side,
    Alignment, Protection, NamedStyle
)
from openpyxl.styles.colors import Color
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import random

OUTPUT_DIR = Path(__file__).parent.parent / "test"


def create_kitchen_sink():
    """Create a comprehensive test file with all formatting features."""
    wb = Workbook()

    # === Sheet 1: Fonts & Colors ===
    ws = wb.active
    ws.title = "Fonts & Colors"
    ws.sheet_properties.tabColor = "FF6B6B"

    # Header
    ws["A1"] = "Font & Color Tests"
    ws["A1"].font = Font(size=16, bold=True)
    ws.merge_cells("A1:E1")

    # Different fonts
    fonts = [
        ("Calibri", 11), ("Arial", 12), ("Times New Roman", 11),
        ("Verdana", 10), ("Courier New", 11), ("Georgia", 11)
    ]
    for i, (name, size) in enumerate(fonts, start=3):
        ws.cell(row=i, column=1, value=f"{name} {size}pt")
        ws.cell(row=i, column=1).font = Font(name=name, size=size)

    # Font styles
    ws["C3"] = "Bold"
    ws["C3"].font = Font(bold=True)
    ws["C4"] = "Italic"
    ws["C4"].font = Font(italic=True)
    ws["C5"] = "Underline"
    ws["C5"].font = Font(underline="single")
    ws["C6"] = "Double Underline"
    ws["C6"].font = Font(underline="double")
    ws["C7"] = "Strikethrough"
    ws["C7"].font = Font(strike=True)
    ws["C8"] = "Bold Italic"
    ws["C8"].font = Font(bold=True, italic=True)

    # Font colors
    colors = [
        ("FF0000", "Red"), ("00FF00", "Green"), ("0000FF", "Blue"),
        ("FF00FF", "Magenta"), ("00FFFF", "Cyan"), ("FFA500", "Orange")
    ]
    for i, (color, name) in enumerate(colors, start=3):
        ws.cell(row=i, column=5, value=name)
        ws.cell(row=i, column=5).font = Font(color=color)

    # Background fills
    ws["A11"] = "Background Fills:"
    ws["A11"].font = Font(bold=True)
    fill_colors = ["FFFF00", "90EE90", "ADD8E6", "FFB6C1", "DDA0DD", "F0E68C"]
    for i, color in enumerate(fill_colors):
        cell = ws.cell(row=12, column=i+1, value=f"Fill {i+1}")
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Pattern fills
    ws["A14"] = "Pattern Fills:"
    ws["A14"].font = Font(bold=True)
    patterns = ["gray125", "gray0625", "darkGray", "mediumGray", "lightGray", "darkHorizontal"]
    for i, pattern in enumerate(patterns):
        cell = ws.cell(row=15, column=i+1, value=pattern[:8])
        cell.fill = PatternFill(start_color="000000", end_color="FFFFFF", fill_type=pattern)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["E"].width = 12

    # === Sheet 2: Borders ===
    ws2 = wb.create_sheet("Borders")
    ws2.sheet_properties.tabColor = "4ECDC4"

    ws2["A1"] = "Border Tests"
    ws2["A1"].font = Font(size=16, bold=True)
    ws2.merge_cells("A1:E1")

    # Border styles
    border_styles = ["thin", "medium", "thick", "double", "dotted", "dashed"]
    for i, style in enumerate(border_styles, start=3):
        cell = ws2.cell(row=i, column=2, value=style)
        side = Side(style=style, color="000000")
        cell.border = Border(left=side, right=side, top=side, bottom=side)

    # Colored borders
    ws2["D3"] = "Red Border"
    ws2["D3"].border = Border(
        left=Side(style="medium", color="FF0000"),
        right=Side(style="medium", color="FF0000"),
        top=Side(style="medium", color="FF0000"),
        bottom=Side(style="medium", color="FF0000")
    )

    ws2["D5"] = "Mixed"
    ws2["D5"].border = Border(
        left=Side(style="thick", color="FF0000"),
        right=Side(style="thick", color="00FF00"),
        top=Side(style="thick", color="0000FF"),
        bottom=Side(style="thick", color="FF00FF")
    )

    # Partial borders
    ws2["D7"] = "Top only"
    ws2["D7"].border = Border(top=Side(style="medium", color="000000"))
    ws2["D8"] = "Bottom only"
    ws2["D8"].border = Border(bottom=Side(style="medium", color="000000"))
    ws2["D9"] = "Left+Right"
    ws2["D9"].border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000")
    )

    ws2.column_dimensions["B"].width = 15
    ws2.column_dimensions["D"].width = 15

    # === Sheet 3: Alignment & Sizing ===
    ws3 = wb.create_sheet("Alignment")
    ws3.sheet_properties.tabColor = "45B7D1"

    ws3["A1"] = "Alignment & Sizing"
    ws3["A1"].font = Font(size=16, bold=True)
    ws3.merge_cells("A1:E1")

    # Horizontal alignment
    h_aligns = ["left", "center", "right", "justify"]
    for i, align in enumerate(h_aligns):
        cell = ws3.cell(row=3, column=i+1, value=f"H: {align}")
        cell.alignment = Alignment(horizontal=align)

    # Vertical alignment
    v_aligns = ["top", "center", "bottom"]
    ws3.row_dimensions[5].height = 40
    for i, align in enumerate(v_aligns):
        cell = ws3.cell(row=5, column=i+1, value=f"V: {align}")
        cell.alignment = Alignment(vertical=align)

    # Text wrap
    ws3["A7"] = "This is a long text that should wrap to multiple lines in the cell"
    ws3["A7"].alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 20
    ws3.row_dimensions[7].height = 45

    # Rotated text
    ws3["C7"] = "45 degrees"
    ws3["C7"].alignment = Alignment(text_rotation=45)
    ws3["D7"] = "90 degrees"
    ws3["D7"].alignment = Alignment(text_rotation=90)
    ws3["E7"] = "-45 degrees"
    ws3["E7"].alignment = Alignment(text_rotation=135)

    # Merged cells
    ws3["A10"] = "This is a merged cell region"
    ws3["A10"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.merge_cells("A10:C12")
    ws3["A10"].fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    # Different row heights
    for i, height in enumerate([15, 25, 35, 45], start=14):
        ws3.row_dimensions[i].height = height
        ws3.cell(row=i, column=1, value=f"Height: {height}pt")

    # Different column widths
    for i, width in enumerate([8, 15, 25, 35], start=1):
        ws3.column_dimensions[get_column_letter(i)].width = width

    # === Sheet 4: Numbers & Dates ===
    ws4 = wb.create_sheet("Numbers")
    ws4.sheet_properties.tabColor = "96CEB4"

    ws4["A1"] = "Number Formats"
    ws4["A1"].font = Font(size=16, bold=True)
    ws4.merge_cells("A1:D1")

    # Various number formats
    formats = [
        (1234.567, "General", "General"),
        (1234.567, "#,##0.00", "Thousands"),
        (0.4567, "0.00%", "Percent"),
        (1234.56, "$#,##0.00", "Currency"),
        (1234.56, '"$"#,##0.00_);[Red]("$"#,##0.00)', "Accounting"),
        (0.5, "# ?/?", "Fraction"),
        (1234567, "0.00E+00", "Scientific"),
        (date(2024, 3, 15), "YYYY-MM-DD", "Date ISO"),
        (date(2024, 3, 15), "MM/DD/YYYY", "Date US"),
        (date(2024, 3, 15), "DD-MMM-YYYY", "Date Long"),
        (datetime(2024, 3, 15, 14, 30, 0), "HH:MM:SS", "Time"),
        (datetime(2024, 3, 15, 14, 30, 0), "YYYY-MM-DD HH:MM", "DateTime"),
    ]

    ws4["A3"] = "Value"
    ws4["B3"] = "Format"
    ws4["C3"] = "Result"
    ws4["A3"].font = ws4["B3"].font = ws4["C3"].font = Font(bold=True)

    for i, (value, fmt, label) in enumerate(formats, start=4):
        ws4.cell(row=i, column=1, value=str(value) if not isinstance(value, (date, datetime)) else value.isoformat())
        ws4.cell(row=i, column=2, value=label)
        cell = ws4.cell(row=i, column=3, value=value)
        cell.number_format = fmt

    # Negative numbers
    ws4["A18"] = "Negative Numbers:"
    ws4["A18"].font = Font(bold=True)
    ws4.cell(row=19, column=1, value=-1234.56).number_format = "#,##0.00"
    ws4.cell(row=19, column=2, value=-1234.56).number_format = "#,##0.00;[Red]-#,##0.00"
    ws4.cell(row=19, column=3, value=-1234.56).number_format = "#,##0.00_);(#,##0.00)"

    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 15
    ws4.column_dimensions["C"].width = 20

    # === Sheet 5: Special Features ===
    ws5 = wb.create_sheet("Special")
    ws5.sheet_properties.tabColor = "FFEAA7"

    ws5["A1"] = "Special Features"
    ws5["A1"].font = Font(size=16, bold=True)
    ws5.merge_cells("A1:D1")

    # Hyperlinks
    ws5["A3"] = "Hyperlinks:"
    ws5["A3"].font = Font(bold=True)
    ws5["A4"] = "Click me!"
    ws5["A4"].hyperlink = "https://example.com"
    ws5["A4"].font = Font(color="0563C1", underline="single")

    # Comments
    ws5["A6"] = "Comments:"
    ws5["A6"].font = Font(bold=True)
    ws5["A7"] = "Hover over me"
    ws5["A7"].comment = Comment("This is a comment!\nWith multiple lines.", "Test Author")

    # Frozen panes
    ws5["A9"] = "This sheet has frozen panes (row 1)"
    ws5.freeze_panes = "A2"

    # Rich text (using multiple cells to simulate)
    ws5["A11"] = "Rich Text (simulated with formatting):"
    ws5["A11"].font = Font(bold=True)

    ws5.column_dimensions["A"].width = 30

    # === Sheet 6: Hidden Sheet (for testing visibility) ===
    ws6 = wb.create_sheet("Hidden Sheet")
    ws6["A1"] = "This sheet is hidden"
    ws6.sheet_state = "hidden"

    # Save
    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / "kitchen_sink.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_large_file(rows=5000, cols=20):
    """Create a large file for load testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Large Dataset"

    # Header row
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c, value=f"Column {c}")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Data rows with mixed formatting
    colors = ["FFFF00", "90EE90", "ADD8E6", "FFB6C1", "FFFFFF"]

    for r in range(2, rows + 2):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)

            # Vary the content type
            if c == 1:
                cell.value = r - 1  # Row number
            elif c == 2:
                cell.value = f"Row {r-1}"
            elif c % 3 == 0:
                cell.value = random.uniform(0, 10000)
                cell.number_format = "#,##0.00"
            elif c % 3 == 1:
                cell.value = random.uniform(0, 1)
                cell.number_format = "0.00%"
            else:
                cell.value = date(2020 + random.randint(0, 4), random.randint(1, 12), random.randint(1, 28))
                cell.number_format = "YYYY-MM-DD"

            # Add some formatting variety
            if r % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            if r % 10 == 0:
                cell.font = Font(bold=True)

        # Progress indicator
        if r % 1000 == 0:
            print(f"  Generated {r}/{rows} rows...")

    # Freeze header row
    ws.freeze_panes = "A2"

    # Set column widths
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12

    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / f"large_{rows}x{cols}.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def create_colors_only():
    """Create a file focused on color testing."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Colors"

    # Rainbow gradient
    rainbow = [
        "FF0000", "FF7F00", "FFFF00", "00FF00",
        "0000FF", "4B0082", "9400D3"
    ]
    for i, color in enumerate(rainbow):
        for j in range(5):
            cell = ws.cell(row=j+1, column=i+1)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Grayscale
    for i in range(10):
        gray = format(int(255 * i / 9), '02X') * 3
        cell = ws.cell(row=7, column=i+1)
        cell.fill = PatternFill(start_color=gray, end_color=gray, fill_type="solid")
        cell.value = f"#{gray}"
        if i < 5:
            cell.font = Font(color="FFFFFF")

    # Theme-like colors (simulating Excel themes)
    theme_colors = [
        ["FFFFFF", "000000", "E7E6E6", "44546A", "4472C4", "ED7D31"],
        ["D0CECE", "7F7F7F", "AEAAAA", "8497B0", "8FAADC", "F4B183"],
        ["A5A5A5", "595959", "757171", "ACB9CA", "B4C6E7", "F8CBAD"],
    ]

    for r, row_colors in enumerate(theme_colors, start=9):
        for c, color in enumerate(row_colors, start=1):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            # Add contrasting text
            luminance = int(color[0:2], 16) * 0.299 + int(color[2:4], 16) * 0.587 + int(color[4:6], 16) * 0.114
            cell.font = Font(color="000000" if luminance > 128 else "FFFFFF")
            cell.value = f"#{color}"

    for c in range(1, 11):
        ws.column_dimensions[get_column_letter(c)].width = 10

    OUTPUT_DIR.mkdir(exist_ok=True)
    output_path = OUTPUT_DIR / "colors_test.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path


def update_manifest():
    """Update manifest.json with all xlsx files in test directory."""
    xlsx_files = sorted(OUTPUT_DIR.glob("*.xlsx"))

    manifest = []
    for f in xlsx_files:
        size = f.stat().st_size
        # Human readable size
        if size > 1_000_000:
            size_str = f"{size / 1_000_000:.1f}MB"
        elif size > 1_000:
            size_str = f"{size / 1_000:.1f}KB"
        else:
            size_str = f"{size}B"

        manifest.append({
            "name": f.name,
            "size": size_str,
        })

    manifest_path = OUTPUT_DIR / "manifest.json"
    with open(manifest_path, "w") as mf:
        json.dump(manifest, mf, indent=2)
    print(f"Updated: {manifest_path} ({len(manifest)} files)")


def main():
    parser = argparse.ArgumentParser(description="Generate test Excel files")
    parser.add_argument("--only", choices=["kitchen", "large", "colors"],
                       help="Generate only specific file")
    parser.add_argument("--rows", type=int, default=5000,
                       help="Number of rows for large file (default: 5000)")
    parser.add_argument("--cols", type=int, default=20,
                       help="Number of columns for large file (default: 20)")
    parser.add_argument("--manifest-only", action="store_true",
                       help="Only update manifest, don't generate files")
    args = parser.parse_args()

    print(f"Output directory: {OUTPUT_DIR}")

    if args.manifest_only:
        update_manifest()
        return

    if args.only == "kitchen":
        create_kitchen_sink()
    elif args.only == "large":
        create_large_file(args.rows, args.cols)
    elif args.only == "colors":
        create_colors_only()
    else:
        # Generate all
        create_kitchen_sink()
        create_colors_only()
        create_large_file(args.rows, args.cols)

    # Always update manifest after generating files
    update_manifest()

    print("\nDone!")


if __name__ == "__main__":
    main()
