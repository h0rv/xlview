# /// script
# requires-python = ">=3.10"
# dependencies = [
#     "openpyxl>=3.1.0",
#     "pillow>=10.0.0",
# ]
# ///
"""
Generate a comprehensive kitchen sink XLSX file with all major features:
- Charts (bar, line, pie, area)
- Embedded images
- Data validation dropdowns
- Conditional formatting (color scales, data bars, icon sets)
- Merged cells
- Frozen panes
- Rich text / styled text
- Comments
- Hyperlinks
- Various cell styles (borders, fills, fonts, alignment)
"""

import io
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection,
    GradientFill, Color
)
from openpyxl.formatting.rule import (
    ColorScaleRule, DataBarRule, IconSetRule, FormulaRule, CellIsRule
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

# Create a simple PNG image in memory (a red square)
def create_test_image(color=(255, 0, 0), size=(100, 100)):
    from PIL import Image as PILImage
    img = PILImage.new('RGB', size, color)
    img_bytes = io.BytesIO()
    img.save(img_bytes, format='PNG')
    img_bytes.seek(0)
    return img_bytes


def create_kitchen_sink():
    wb = Workbook()

    # Define common fonts
    black_font = Font(color="000000")
    white_font = Font(color="FFFFFF")

    # =========================================================================
    # Sheet 1: Charts
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Charts"
    ws1.sheet_properties.tabColor = "4472C4"  # Blue tab

    # Sample data for charts
    ws1['A1'] = "Category"
    ws1['B1'] = "Series 1"
    ws1['C1'] = "Series 2"
    ws1['D1'] = "Series 3"

    categories = ["Q1", "Q2", "Q3", "Q4"]
    data1 = [10, 25, 15, 30]
    data2 = [20, 15, 25, 20]
    data3 = [15, 30, 20, 25]

    # Default black font for data cells
    black_font = Font(color="000000")

    for i, (cat, d1, d2, d3) in enumerate(zip(categories, data1, data2, data3), 2):
        ws1[f'A{i}'] = cat
        ws1[f'A{i}'].font = black_font
        ws1[f'B{i}'] = d1
        ws1[f'B{i}'].font = black_font
        ws1[f'C{i}'] = d2
        ws1[f'C{i}'].font = black_font
        ws1[f'D{i}'] = d3
        ws1[f'D{i}'].font = black_font

    # Style the header row
    for col in ['A', 'B', 'C', 'D']:
        cell = ws1[f'{col}1']
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Quarterly Sales"
    bar_chart.y_axis.title = "Amount"
    bar_chart.x_axis.title = "Quarter"

    data_ref = Reference(ws1, min_col=2, min_row=1, max_col=4, max_row=5)
    cats_ref = Reference(ws1, min_col=1, min_row=2, max_row=5)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(cats_ref)
    bar_chart.shape = 4
    ws1.add_chart(bar_chart, "F2")

    # Line Chart
    line_chart = LineChart()
    line_chart.title = "Sales Trend"
    line_chart.y_axis.title = "Amount"
    line_chart.x_axis.title = "Quarter"
    line_chart.add_data(data_ref, titles_from_data=True)
    line_chart.set_categories(cats_ref)
    ws1.add_chart(line_chart, "F18")

    # Pie Chart data
    ws1['G1'] = "Product"
    ws1['G1'].font = black_font
    ws1['H1'] = "Sales"
    ws1['H1'].font = black_font
    pie_data = [("Widgets", 35), ("Gadgets", 25), ("Gizmos", 20), ("Things", 20)]
    for i, (prod, sales) in enumerate(pie_data, 2):
        ws1[f'G{i}'] = prod
        ws1[f'G{i}'].font = black_font
        ws1[f'H{i}'] = sales
        ws1[f'H{i}'].font = black_font

    pie_chart = PieChart()
    pie_chart.title = "Product Distribution"
    pie_ref = Reference(ws1, min_col=8, min_row=2, max_row=5)
    pie_cats = Reference(ws1, min_col=7, min_row=2, max_row=5)
    pie_chart.add_data(pie_ref)
    pie_chart.set_categories(pie_cats)
    ws1.add_chart(pie_chart, "P2")

    # Area Chart
    area_chart = AreaChart()
    area_chart.title = "Cumulative Sales"
    area_chart.add_data(data_ref, titles_from_data=True)
    area_chart.set_categories(cats_ref)
    ws1.add_chart(area_chart, "P18")

    # =========================================================================
    # Sheet 2: Images
    # =========================================================================
    ws2 = wb.create_sheet("Images")
    ws2.sheet_properties.tabColor = "70AD47"  # Green tab

    ws2['A1'] = "This sheet contains embedded images"
    ws2['A1'].font = Font(bold=True, size=14, color="000000")

    # Create and add test images
    red_img = Image(create_test_image((255, 0, 0), (80, 80)))
    red_img.anchor = "B3"
    ws2.add_image(red_img)
    ws2['B2'] = "Red"
    ws2['B2'].font = black_font

    green_img = Image(create_test_image((0, 255, 0), (80, 80)))
    green_img.anchor = "D3"
    ws2.add_image(green_img)
    ws2['D2'] = "Green"
    ws2['D2'].font = black_font

    blue_img = Image(create_test_image((0, 0, 255), (80, 80)))
    blue_img.anchor = "F3"
    ws2.add_image(blue_img)
    ws2['F2'] = "Blue"
    ws2['F2'].font = black_font

    yellow_img = Image(create_test_image((255, 255, 0), (80, 80)))
    yellow_img.anchor = "H3"
    ws2.add_image(yellow_img)
    ws2['H2'] = "Yellow"
    ws2['H2'].font = black_font

    # =========================================================================
    # Sheet 3: Data Validation
    # =========================================================================
    ws3 = wb.create_sheet("Data Validation")
    ws3.sheet_properties.tabColor = "FFC000"  # Orange tab

    ws3['A1'] = "Dropdown Examples"
    ws3['A1'].font = Font(bold=True, size=14)

    # List validation (dropdown)
    ws3['A3'] = "Select Status:"
    dv_status = DataValidation(
        type="list",
        formula1='"Active,Pending,Completed,Cancelled"',
        allow_blank=True,
        showDropDown=False  # False means SHOW the dropdown arrow
    )
    dv_status.prompt = "Select a status"
    dv_status.promptTitle = "Status"
    ws3.add_data_validation(dv_status)
    dv_status.add('B3')
    ws3['B3'] = "Active"

    # Priority dropdown
    ws3['A4'] = "Select Priority:"
    dv_priority = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        showDropDown=False
    )
    ws3.add_data_validation(dv_priority)
    dv_priority.add('B4')
    ws3['B4'] = "Medium"

    # Number range validation
    ws3['A6'] = "Enter Age (1-120):"
    dv_age = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="120"
    )
    dv_age.error = "Age must be between 1 and 120"
    dv_age.errorTitle = "Invalid Age"
    ws3.add_data_validation(dv_age)
    dv_age.add('B6')
    ws3['B6'] = 25

    # Date validation
    ws3['A7'] = "Enter Date (2024+):"
    dv_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="2024-01-01"
    )
    ws3.add_data_validation(dv_date)
    dv_date.add('B7')

    # Yes/No dropdown for multiple cells
    ws3['A9'] = "Completed?"
    ws3['A10'] = "Approved?"
    ws3['A11'] = "Verified?"
    dv_yesno = DataValidation(
        type="list",
        formula1='"Yes,No"',
        showDropDown=False
    )
    ws3.add_data_validation(dv_yesno)
    dv_yesno.add('B9:B11')
    ws3['B9'] = "Yes"
    ws3['B10'] = "No"
    ws3['B11'] = "Yes"

    # =========================================================================
    # Sheet 4: Conditional Formatting
    # =========================================================================
    ws4 = wb.create_sheet("Conditional Formatting")
    ws4.sheet_properties.tabColor = "ED7D31"  # Orange tab

    ws4['A1'] = "Conditional Formatting Examples"
    ws4['A1'].font = Font(bold=True, size=14)
    ws4.merge_cells('A1:E1')

    # Color Scale (2-color: red to green)
    ws4['A3'] = "Color Scale (2-color)"
    ws4['A3'].font = Font(bold=True)
    for i, val in enumerate([10, 30, 50, 70, 90], 4):
        ws4[f'A{i}'] = val

    ws4.conditional_formatting.add(
        'A4:A8',
        ColorScaleRule(
            start_type='min', start_color='FF0000',
            end_type='max', end_color='00FF00'
        )
    )

    # Color Scale (3-color)
    ws4['B3'] = "Color Scale (3-color)"
    ws4['B3'].font = Font(bold=True)
    for i, val in enumerate([0, 25, 50, 75, 100], 4):
        ws4[f'B{i}'] = val

    ws4.conditional_formatting.add(
        'B4:B8',
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )

    # Data Bars
    ws4['C3'] = "Data Bars"
    ws4['C3'].font = Font(bold=True)
    for i, val in enumerate([20, 40, 60, 80, 100], 4):
        ws4[f'C{i}'] = val

    ws4.conditional_formatting.add(
        'C4:C8',
        DataBarRule(
            start_type='min',
            end_type='max',
            color='4472C4',
            showValue=True,
            minLength=None,
            maxLength=None
        )
    )

    # Icon Sets (3 arrows)
    ws4['D3'] = "Icon Set (Arrows)"
    ws4['D3'].font = Font(bold=True)
    for i, val in enumerate([15, 35, 55, 75, 95], 4):
        ws4[f'D{i}'] = val

    ws4.conditional_formatting.add(
        'D4:D8',
        IconSetRule(
            icon_style='3Arrows',
            type='percent',
            values=[0, 33, 67]
        )
    )

    # Icon Sets (traffic lights)
    ws4['E3'] = "Icon Set (Lights)"
    ws4['E3'].font = Font(bold=True)
    for i, val in enumerate([1, 2, 3, 2, 1], 4):
        ws4[f'E{i}'] = val

    ws4.conditional_formatting.add(
        'E4:E8',
        IconSetRule(
            icon_style='3TrafficLights1',
            type='num',
            values=[0, 2, 3]
        )
    )

    # Cell-based rules
    ws4['A10'] = "Cell Rules Examples"
    ws4['A10'].font = Font(bold=True)
    ws4.merge_cells('A10:E10')

    ws4['A11'] = "Value"
    ws4['B11'] = ">50 (Green)"
    ws4['C11'] = "<30 (Red)"

    for i, val in enumerate([10, 40, 60, 25, 80], 12):
        ws4[f'A{i}'] = val
        ws4[f'B{i}'] = val
        ws4[f'C{i}'] = val

    # Greater than 50 = green fill
    ws4.conditional_formatting.add(
        'B12:B16',
        CellIsRule(
            operator='greaterThan',
            formula=['50'],
            fill=PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        )
    )

    # Less than 30 = red fill
    ws4.conditional_formatting.add(
        'C12:C16',
        CellIsRule(
            operator='lessThan',
            formula=['30'],
            fill=PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        )
    )

    # =========================================================================
    # Sheet 5: Styles & Formatting
    # =========================================================================
    ws5 = wb.create_sheet("Styles")
    ws5.sheet_properties.tabColor = "7030A0"  # Purple tab

    # Freeze panes
    ws5.freeze_panes = 'B2'

    ws5['A1'] = "Style Examples"
    ws5['A1'].font = Font(bold=True, size=16, color="4472C4")

    # Font styles
    ws5['A3'] = "Font Styles:"
    ws5['A3'].font = Font(bold=True)

    ws5['B3'] = "Bold"
    ws5['B3'].font = Font(bold=True)

    ws5['C3'] = "Italic"
    ws5['C3'].font = Font(italic=True)

    ws5['D3'] = "Underline"
    ws5['D3'].font = Font(underline='single')

    ws5['E3'] = "Strikethrough"
    ws5['E3'].font = Font(strike=True)

    ws5['F3'] = "Red Text"
    ws5['F3'].font = Font(color="FF0000")

    ws5['G3'] = "Large"
    ws5['G3'].font = Font(size=18)

    # Fill patterns
    ws5['A5'] = "Fill Patterns:"
    ws5['A5'].font = Font(bold=True)

    ws5['B5'] = "Solid"
    ws5['B5'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    ws5['C5'] = "Gray125"
    ws5['C5'].fill = PatternFill(start_color="000000", end_color="FFFFFF", fill_type="gray125")

    ws5['D5'] = "Gradient"
    ws5['D5'].fill = GradientFill(stop=["FF0000", "0000FF"])

    # Borders
    ws5['A7'] = "Borders:"
    ws5['A7'].font = Font(bold=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    thick_border = Border(
        left=Side(style='thick', color='FF0000'),
        right=Side(style='thick', color='FF0000'),
        top=Side(style='thick', color='FF0000'),
        bottom=Side(style='thick', color='FF0000')
    )

    dashed_border = Border(
        left=Side(style='dashed'),
        right=Side(style='dashed'),
        top=Side(style='dashed'),
        bottom=Side(style='dashed')
    )

    ws5['B7'] = "Thin"
    ws5['B7'].border = thin_border

    ws5['C7'] = "Thick Red"
    ws5['C7'].border = thick_border

    ws5['D7'] = "Dashed"
    ws5['D7'].border = dashed_border

    # Alignment
    ws5['A9'] = "Alignment:"
    ws5['A9'].font = Font(bold=True)

    ws5['B9'] = "Center"
    ws5['B9'].alignment = Alignment(horizontal='center', vertical='center')

    ws5['C9'] = "Right"
    ws5['C9'].alignment = Alignment(horizontal='right')

    ws5['D9'] = "Wrapped Long Text That Wraps"
    ws5['D9'].alignment = Alignment(wrap_text=True)
    ws5.column_dimensions['D'].width = 15

    ws5['E9'] = "45 Degrees"
    ws5['E9'].alignment = Alignment(text_rotation=45)

    ws5['F9'] = "Indent"
    ws5['F9'].alignment = Alignment(indent=2)

    # Merged cells
    ws5['A11'] = "Merged Cells:"
    ws5['A11'].font = Font(bold=True)

    ws5.merge_cells('B11:D11')
    ws5['B11'] = "This is merged across 3 cells"
    ws5['B11'].alignment = Alignment(horizontal='center')
    ws5['B11'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws5.merge_cells('B12:B14')
    ws5['B12'] = "Vertical merge"
    ws5['B12'].alignment = Alignment(horizontal='center', vertical='center')
    ws5['B12'].fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # =========================================================================
    # Sheet 6: Comments & Hyperlinks
    # =========================================================================
    ws6 = wb.create_sheet("Comments & Links")
    ws6.sheet_properties.tabColor = "00B0F0"  # Light blue tab

    ws6['A1'] = "Comments and Hyperlinks"
    ws6['A1'].font = Font(bold=True, size=14)

    # Comments
    ws6['A3'] = "Cells with comments:"
    ws6['A3'].font = Font(bold=True)

    ws6['B3'] = "Hover here"
    ws6['B3'].comment = Comment("This is a comment!\nIt can have multiple lines.", "Author")

    ws6['C3'] = "Another comment"
    ws6['C3'].comment = Comment("Important note about this cell.", "Reviewer")

    # Hyperlinks
    ws6['A5'] = "Hyperlinks:"
    ws6['A5'].font = Font(bold=True)

    ws6['B5'] = "Google"
    ws6['B5'].hyperlink = "https://www.google.com"
    ws6['B5'].font = Font(color="0563C1", underline='single')

    ws6['C5'] = "Microsoft"
    ws6['C5'].hyperlink = "https://www.microsoft.com"
    ws6['C5'].font = Font(color="0563C1", underline='single')

    ws6['B6'] = "Email Link"
    ws6['B6'].hyperlink = "mailto:test@example.com"
    ws6['B6'].font = Font(color="0563C1", underline='single')

    # Internal link
    ws6['B7'] = "Go to Charts Sheet"
    ws6['B7'].hyperlink = "#Charts!A1"
    ws6['B7'].font = Font(color="0563C1", underline='single')

    # =========================================================================
    # Sheet 7: Numbers & Dates
    # =========================================================================
    ws7 = wb.create_sheet("Numbers & Dates")
    ws7.sheet_properties.tabColor = "FF6B6B"  # Red tab

    ws7['A1'] = "Number Formats"
    ws7['A1'].font = Font(bold=True, size=14)

    # Various number formats
    ws7['A3'] = "Format"
    ws7['B3'] = "Value"
    ws7['A3'].font = Font(bold=True)
    ws7['B3'].font = Font(bold=True)

    ws7['A4'] = "General"
    ws7['B4'] = 1234.5678

    ws7['A5'] = "Currency"
    ws7['B5'] = 1234.56
    ws7['B5'].number_format = '$#,##0.00'

    ws7['A6'] = "Percentage"
    ws7['B6'] = 0.756
    ws7['B6'].number_format = '0.00%'

    ws7['A7'] = "Scientific"
    ws7['B7'] = 123456789
    ws7['B7'].number_format = '0.00E+00'

    ws7['A8'] = "Date"
    from datetime import datetime
    ws7['B8'] = datetime(2024, 6, 15)
    ws7['B8'].number_format = 'YYYY-MM-DD'

    ws7['A9'] = "Time"
    ws7['B9'] = datetime(2024, 1, 1, 14, 30, 45)
    ws7['B9'].number_format = 'HH:MM:SS'

    ws7['A10'] = "DateTime"
    ws7['B10'] = datetime(2024, 12, 25, 10, 30)
    ws7['B10'].number_format = 'YYYY-MM-DD HH:MM'

    ws7['A11'] = "Accounting"
    ws7['B11'] = -1234.56
    ws7['B11'].number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'

    ws7['A12'] = "Fraction"
    ws7['B12'] = 0.5
    ws7['B12'].number_format = '# ?/?'

    # =========================================================================
    # Adjust column widths
    # =========================================================================
    for ws in wb.worksheets:
        for col in range(1, 20):
            letter = get_column_letter(col)
            if ws.column_dimensions[letter].width is None:
                ws.column_dimensions[letter].width = 15

    # Ensure all cells have explicit black text color (unless white is intended)
    # This fixes issues where openpyxl defaults to theme colors that may appear white
    white_bg_cells = {"#FFFFFF", "#ffffff", "FFFFFF", "ffffff", None}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check if cell has a fill that makes white text appropriate
                    has_dark_bg = False
                    try:
                        if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                            fg = cell.fill.fgColor
                            if fg.rgb and fg.rgb not in white_bg_cells:
                                # Has a colored background - check if dark
                                rgb = fg.rgb[-6:] if len(fg.rgb) >= 6 else fg.rgb
                                try:
                                    r = int(rgb[0:2], 16)
                                    g = int(rgb[2:4], 16)
                                    b = int(rgb[4:6], 16)
                                    # If luminance is low, it's a dark background
                                    luminance = 0.299 * r + 0.587 * g + 0.114 * b
                                    has_dark_bg = luminance < 128
                                except (ValueError, IndexError):
                                    pass
                    except AttributeError:
                        pass  # GradientFill or other fill types without fgColor

                    # If font has no color or has white color on light bg, set to black
                    current_font = cell.font
                    if current_font:
                        font_color = current_font.color
                        if font_color is None or (font_color.rgb in ("FFFFFFFF", "00FFFFFF", "FFFFFF") and not has_dark_bg):
                            # Create new font with black color, preserving other attributes
                            cell.font = Font(
                                name=current_font.name,
                                size=current_font.size,
                                bold=current_font.bold,
                                italic=current_font.italic,
                                underline=current_font.underline,
                                strike=current_font.strike,
                                color="000000"
                            )

    # Save
    output_path = "/Users/robby/projects/xlview/test/kitchen_sink_v2.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

    # Show summary
    print("\nSheets created:")
    for ws in wb.worksheets:
        print(f"  - {ws.title} (tab color: {ws.sheet_properties.tabColor})")

    print("\nFeatures included:")
    print("  - 4 chart types (bar, line, pie, area)")
    print("  - 4 embedded images")
    print("  - Data validation dropdowns")
    print("  - Conditional formatting (color scales, data bars, icon sets)")
    print("  - Various cell styles (fonts, fills, borders, alignment)")
    print("  - Merged cells")
    print("  - Frozen panes")
    print("  - Comments")
    print("  - Hyperlinks")
    print("  - Number formats")


if __name__ == "__main__":
    create_kitchen_sink()
